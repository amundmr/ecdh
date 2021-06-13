from utils import *

def read_csv(filepath):
    import numpy as np
    #Open file
    with open(filepath, 'r', encoding = "ISO-8859-1") as f:
        #Read the first 2 lines which doesnt contain any data
        for i in range(2):
            f.readline()

        #Find what columns have our data
        Labels = f.readline().split(",") #Save the line of labels of which we are interested in
        V_col = 0
        mAh_col = 0
        for i,label in enumerate(Labels):
            if "Voltage(V)" in label:
                V_col = i
            if "Capacity(mAh)" in label:
                mAh_col = i

        #Read rest of data as list of lines
        data = f.readlines()
    
    #Index where in the file charges and discharges occur
    Chg_indxs = []
    DChg_indxs = []

    for i,line in enumerate(data):
        line_split = line.split(",")
        if "CC_Chg" in line_split[2]:
            Chg_indxs.append(i)
        elif "CC_DChg" in line_split[2]:
            DChg_indxs.append(i)

    charges = []
    discharges = []
    #Returns two arrays with Voltage and Capacity for the selected indexes of data.
    def slice_data(start, end, data):
        if end == -1:
            end = len(data)

        #Spawning data arrays
        start +=1
        end -=2
        voltages = np.zeros(end-start)
        capacities = np.zeros_like(voltages)
        for j,line in enumerate(data[start: end]):#Taking the slice of the data arr
            line = line.split(",")
            voltages[j] = float(line[V_col])
            capacities[j] = float(line[mAh_col])
        return voltages, capacities

    #Create one big list of all indexes in order to know which one is the next
    all_indxs = Chg_indxs + DChg_indxs
    all_indxs.sort()
    # Iterating over all indexes of charge start
    for i, indx in enumerate(Chg_indxs):
        i_all = all_indxs.index(indx) #Finding the corresponding item in list of all indexes
        start = indx   #Setting start of Charge to the charge index
        try:            #Setting end of charge to discharge index, or last index of data.
            end = all_indxs[i_all+1]
        except:
            end = len(data)-1
        #Getting the data slice, and putting it in charges list.
        voltages, capacities = slice_data(start, end, data)
        #print(voltages)
        charges.append((voltages, capacities))
    
    # Iterating over all indexes of discharge start
    for i, indx in enumerate(DChg_indxs):
        i_all = all_indxs.index(indx) #Finding the corresponding item in list of all indexes
        start = indx   #Setting start of Discharge to the discharge index
        try:            #Setting end of charge to charge index, or last index of data.
            end = all_indxs[i_all+1]
        except:
            end = len(data)-1
        #Getting the data slice, and putting it in charges list.
        voltages, capacities = slice_data(start, end, data)
        discharges.append((voltages, capacities))

            
    return charges, discharges


def read_xlsx(filename):
    import openpyxl as px
    import numpy as np

    # Open workbook
    wb = px.load_workbook(filename)
    # Get sheet
    sheet = wb[wb.sheetnames[3]]
    max_row = sheet.max_row-1 #-1 removes the first row with text
    # Spawn arrays for all raw data of columns
    status = []
    voltage = np.zeros(max_row)
    capacity = np.zeros(max_row)

    # Insert all data in arrays
    for i in range(max_row):
        voltage[i] = sheet.cell(row=i+2, column = 7).value
        capacity[i] = sheet.cell(row=i+2, column = 8).value
        status.append(sheet.cell(row=i+2, column=2).value)

    # Find where a charge and a discharge starts. Since charging occurs first, charges will be at even indexes of indx, eg: 0, 2, 4..
    indx =[] # Just a list to store where in the raw data a charge (even index) or discharge (odd index)
    for i in range(1,len(status)):
        if status[i] == "CC Chg" and status[i-1] != "CC Chg": # If status just changed to CC Chg, this is a charge start
            indx.append(i)
        elif status[i] == "CC DChg" and status[i-1] != "CC DChg": # If status just changed to CC DChg, this is a discharge start
            indx.append(i)
    indx.append(max_row) #Inserting last elem means we catch the end cycle!

    # Making lists for the cycles
    charges = []
    discharges = []

    # Inserting slices of raw data (corresponding to 1 cycle) into the list of charges
    for i in range(len(indx)-1): #-1 is to not run out of indexes when i+1
        if (i % 2) == 0: #index is even -> charge cycle
            charges.append((voltage[indx[i]:indx[i+1]], capacity[indx[i]:indx[i+1]])) #Inserting tuple of two arrays sliced to be one cycle of voltages and capacities.
        else: #Not even? Then its a discharge cycle
            discharges.append((voltage[indx[i]:indx[i+1]], capacity[indx[i]:indx[i+1]]))

    return charges, discharges


def mpt_biologic_to_vq_old(filepath):
    import numpy as np
    f = open(filepath, 'r') #open the filepath for the mpt file
    # now we skip all the data in the start and jump straight to the dense data
    headerlines = 0
    for line in f:
        if "Nb header lines :" in line:
            headerlines = int(line.split(':')[-1])
            break #breaks for loop when headerlines is found
    
    for line in f:
        if "Characteristic mass :" in line:
            active_mass = float(line.split(':')[-1][:-3].replace(',', '.'))/1000
            print("Active mass found in file to be: " + str(active_mass) + "g")
            break #breaks loop when active mass is found

    #Skip all headerlines
    for i in range(headerlines):
        f.readline()


    #Put all the data in one big list of lines
    data = f.readlines()

    #Spawning massive data arrays for Potential, Charge and Cycle num
    Ewe_arr = np.zeros(len(data))
    Q_chg_dischg_arr = np.zeros_like(Ewe_arr)
    Cycle_arr = np.zeros_like(Ewe_arr)
    ox_red = np.zeros_like(Ewe_arr)

    #Putting all data (potential, capacity, cyclenum) in arrays
    i=-1
    for line in data: #Looping over each line in the dense data
        
        i+=1
        line_lst = line.split() #Splitting the line to a list
        Cycle_arr[i] = int(float(line_lst[-2].replace(',', '.'))) #The integer of the last element on the line as cycle added to array of cycles
        Ewe_arr[i] = float(line_lst[11].replace(',', '.')) #line elem 11 added to Ewe arr as float
        Q_chg_dischg_arr[i] = float(line_lst[19].replace(',', '.')) / active_mass
        ox_red[i] = int(float(line_lst[1].replace(',', '.')))
        

    #Creating list with indexes of new cycle starts
    list_of_cycleindexes = [0]
    list_of_reversalindexes = []
    for i in range(len(Cycle_arr)-1):
        if Cycle_arr[i]+1 == Cycle_arr[i+1]: #Adding the index where a new cycle starts (stop discharge, start charge)
            list_of_cycleindexes.append(i+1)
        if ox_red[i] == ox_red[i+1]+1: #Adding the index where the charge stops and the discharge starts
            list_of_reversalindexes.append(i+1)

    charges = []
    discharges = []
    indx = list_of_cycleindexes
    [indx.append(i) for i in list_of_reversalindexes]
    indx.sort()

    for i in range(len(indx) - 1):
        if (i%2) == 0: #even index: charge cycle
            charges.append((Ewe_arr[indx[i]:indx[i+1]], Q_chg_dischg_arr[indx[i]:indx[i+1]])) #Inserting tuple of two arrays sliced to be one cycle of voltages and capacities.
        else: #then discharge
            discharges.append((Ewe_arr[indx[i]:indx[i+1]], Q_chg_dischg_arr[indx[i]:indx[i+1]]))
    return charges, discharges
    

def mpt_biologic_to_vq(filepath):
    import numpy as np
    with open(filepath, 'r') as f:#open the filepath for the mpt file
        lines = f.readlines()

    # now we skip all the data in the start and jump straight to the dense data
    headerlines = 0
    for line in lines:
        if "Nb header lines :" in line:
            headerlines = int(line.split(':')[-1])
            break #breaks for loop when headerlines is found
    
    for line in lines:
        if "Characteristic mass :" in line:
            active_mass = float(line.split(':')[-1][:-3].replace(',', '.'))/1000
            print("Active mass found in file to be: " + str(active_mass) + "g")
            break #breaks loop when active mass is found

    #Ox/red-col = 1
    #Ewe-col = 9
    #Cap-col = 17
    #Cyc-col = -2
    #Cur-col = 10
    #time-col = 7
    charges = []
    discharges = []
    newox = False
    oldox = eval(lines[headerlines].split()[1])
    oldcap = 0
    t_prev = 0
    t_cycstart = 0
    tmp_cyc_E = []
    tmp_cyc_C = []
    Cap_cum = []
    for line in lines[headerlines:]:
        if oldox == eval(line.split()[1]):
            tmp_cyc_E.append(eval(line.split()[9]))
            tmp_cyc_C.append(eval(line.split()[17]))
            
            #Manual capacity calculation
            I_cur = eval(line.split()[10])
            t_cur = eval(line.split()[7])/3600 - t_cycstart
            timedelta = t_cur - t_prev
            Cap_cum.append(abs(oldcap + I_cur * timedelta))
            t_prev = t_cur
            oldcap += I_cur * timedelta

        else:
            print("End-Halfcycle-triggered")
            if oldox == 1: #Charge cycle
                charges.append((np.array(tmp_cyc_E), np.array(Cap_cum))) # Making the lists into arrays and putting them in a tuple as a charge cycle
            elif oldox == 0: #Discharge cycle
                discharges.append((np.array(tmp_cyc_E), np.array(Cap_cum)))
            
            # Resetting the temporary lists
            tmp_cyc_E = []
            tmp_cyc_C = []
            Cap_cum = []
            t_prev = 0
            oldcap = 0
            # Setting the new ox status
            oldox = eval(line.split()[1])
            # Must remember to add this line's values!
            tmp_cyc_E.append(eval(line.split()[9]))
            tmp_cyc_C.append(eval(line.split()[17]))

            t_cycstart = eval(line.split()[7])/3600 # Resetting time of start of cycle
            #Manual capacity calculation
            I_cur = eval(line.split()[10])
            t_cur = eval(line.split()[7])/3600 - t_cycstart
            timedelta = t_cur - t_prev
            Cap_cum.append(oldcap + I_cur * timedelta)
            t_prev = t_cur
            oldcap += I_cur * timedelta

    # Adding the last data to arrays if the file ends
    if oldox == 1: #Charge cycle
        print("EOF- with a Charge")
        charges.append((np.array(tmp_cyc_E), np.array(Cap_cum))) # Making the lists into arrays and putting them in a tuple as a charge cycle
    elif oldox == 0: #Discharge cycle
        print("EOF- with a Discharge")
        discharges.append((np.array(tmp_cyc_E), np.array(Cap_cum)))

    return charges, discharges  


def custom_EC_export(filepath):
    import numpy as np
    with open(filepath, 'r') as f: #open the filepath for the mpt file
        f.readline()
        lines = f.readlines()

    #Ox/red-col = 1
    #Ewe-col = 4
    #Cap-col = 5
    #Cyc-col = 0
    #Cur-col = 3
    #time-col = 2
    charges = []
    discharges = []
    newox = False
    oldox = eval(lines[0].split()[1])
    oldcap = 0
    t_prev = 0
    t_cycstart = 0
    tmp_cyc_E = []
    tmp_cyc_C = []
    Cap_cum = []
    for line in lines:
        if oldox == eval(line.split()[1]):
            tmp_cyc_E.append(eval(line.split()[4]))
            tmp_cyc_C.append(eval(line.split()[5]))
            
            #Manual capacity calculation
            I_cur = eval(line.split()[3])
            t_cur = eval(line.split()[2])/3600 - t_cycstart
            timedelta = t_cur - t_prev
            Cap_cum.append(abs(oldcap + I_cur * timedelta))
            t_prev = t_cur
            oldcap += I_cur * timedelta

        else:
            #print("End-Halfcycle-triggered")
            if oldox == 1: #Charge cycle
                charges.append((np.array(tmp_cyc_E), np.array(Cap_cum))) # Making the lists into arrays and putting them in a tuple as a charge cycle
            elif oldox == 0: #Discharge cycle
                discharges.append((np.array(tmp_cyc_E), np.array(Cap_cum)))
            
            # Resetting the temporary lists
            tmp_cyc_E = []
            tmp_cyc_C = []
            Cap_cum = []
            t_prev = 0
            oldcap = 0
            # Setting the new ox status
            oldox = eval(line.split()[1])
            # Must remember to add this line's values!
            tmp_cyc_E.append(eval(line.split()[4]))
            tmp_cyc_C.append(eval(line.split()[5]))

            t_cycstart = eval(line.split()[2])/3600 # Resetting time of start of cycle
            #Manual capacity calculation
            I_cur = eval(line.split()[3])
            t_cur = eval(line.split()[2])/3600 - t_cycstart
            timedelta = t_cur - t_prev
            Cap_cum.append(oldcap + I_cur * timedelta)
            t_prev = t_cur
            oldcap += I_cur * timedelta

    # Adding the last data to arrays if the file ends
    if oldox == 1: #Charge cycle
        info("Datafile ended with a Charge")
        charges.append((np.array(tmp_cyc_E), np.array(Cap_cum))) # Making the lists into arrays and putting them in a tuple as a charge cycle
    elif oldox == 0: #Discharge cycle
        info("Datafile ended with a Discharge")
        discharges.append((np.array(tmp_cyc_E), np.array(Cap_cum)))

    return charges, discharges  

def dat_batsmall_to_vq(filename):
    import numpy as np
    data = []
    decode_errors = 0
    decode_error_str = ""
    with open(filename, "r") as f:
        # Skip all non-data lines
        while True:
            line = f.readline()
            if '"V";I:"A";C:"Ah/kg";7' in line:
                break
        # Adding the rest of the readable file to a data array
        while True:
            try:
                line = f.readline()
                if line == '':
                    break
                else:
                    data.append(line)
            except UnicodeDecodeError as e:
                decode_errors += 1
                decode_error_str = e
                continue
        
        if decode_errors > 0:
            warn("Found %i Unicode Decode errors, thus %i lines of data has been missed. Consider getting a safer method of acquiring data. \nComplete error message: " %(decode_errors, decode_errors) + str(decode_error_str))


    charges = []
    discharges = []
    charge = True
    voltages = []
    capacities = []

    for line in data[:-1]:
        #print(line.split(";"))
        try:
            v = float(eval(line.split(";")[1]))
            q = abs(float(eval(line.split(";")[3])))
            voltages.append(v)
            capacities.append(q)
        except:
            if '"V";I:"A";C:"Ah/kg"' in line and charge == True:
                print("Charge end at: V: %.5f, Q: %.5f" %(v,q))
                charges.append((np.array(voltages), np.array(capacities)))
                charge = False
                voltages = []
                capacities = []
            elif '"V";I:"A";C:"Ah/kg"' in line and charge == False:
                print("Discharge end at: V: %.5f, Q: %.5f" %(v,q))
                discharges.append((np.array(voltages), np.array(capacities)))
                charge = True
                voltages = []
                capacities = []

        
    
    #print(charges)
    #print(discharges)
    return charges, discharges