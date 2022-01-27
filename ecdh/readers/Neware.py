# -*- coding: utf-8 -*-
"""Data-readers for Neware"""
import pandas as pd
import numpy as np      
import gc
from ecdh.log import LOG
LOG.error("MATS IS WORKING ON MOVING OVER THE *** TO PD DATAFRAMES")
def read_csv(filepath):
    
    """
    Reads a .csv file from Neware general report
    
    .csv format:
    
    """
    LOG.debug(f"Reading Neware CSV: '{filepath}'")

    # Open file
    with open(filepath, 'r') as f:
        lines = f.readlines()

    # Read all data to pandas dataframe
    big_df = pd.read_csv(filepath, header = 2, delim_whitespace=True)
    pd.DataFrame.fillna(big_df, method='ffill', inplace = True)
    #print(big_df["Unnamed: 0"].head)
    print(big_df.columns)
    print(big_df.head)
    #df = big_df.loc[:,('\tTime(h:min:s.ms)', '\tVoltage(V)', '\tCurrent(mA)', 'Unnamed: 0', '\tCapacity Density(mAh/g)')]
    del big_df #deletes the dataframe
    gc.collect() #Clean unused memory (which is the dataframe above)
    #print(df.head)
    #df.columns = ['time/s','Ewe/V', '<I>/mA', 'cycle number', 'capacity/mAhg'] 
    #print(df.columns)
    #print(df['time/s'].iloc[1])
    def _time_parser(string):
        """Takes neware string of time and returns float in seconds"""
        print(string)
        string = string[2:]
        nums = string.split(":")
        hours = float(nums[0])
        mins = float(nums[1])
        secs = float(nums[2])
        return hours*3600 + mins*60 + secs

    df['time/s'] = df['time/s'].apply(lambda x: _time_parser(x)) #Converting from h to s
    print(df.head)
    """import numpy as np
    #Open file
    with open(filepath, 'r', encoding = "ISO-8859-1") as f:
        #Read the first 2 lines which doesnt contain any data
        for i in range(2):
            f.readline()

        #Find what columns have our data
        Labels = f.readline().split(",") #Save the line of labels of which we are interested in
        V_col = 0
        mAh_col = 0
        for i,label in enumerate(Labels):
            if "Voltage(V)" in label:
                V_col = i
            if "Capacity(mAh)" in label:
                mAh_col = i

        #Read rest of data as list of lines
        data = f.readlines()
    
    #Index where in the file charges and discharges occur
    Chg_indxs = []
    DChg_indxs = []

    for i,line in enumerate(data):
        line_split = line.split(",")
        if "CC_Chg" in line_split[2]:
            Chg_indxs.append(i)
        elif "CC_DChg" in line_split[2]:
            DChg_indxs.append(i)

    charges = []
    discharges = []
    #Returns two arrays with Voltage and Capacity for the selected indexes of data.
    def slice_data(start, end, data):
        if end == -1:
            end = len(data)

        #Spawning data arrays
        start +=1
        end -=2
        voltages = np.zeros(end-start)
        capacities = np.zeros_like(voltages)
        for j,line in enumerate(data[start: end]):#Taking the slice of the data arr
            line = line.split(",")
            voltages[j] = float(line[V_col])
            capacities[j] = float(line[mAh_col])
        return voltages, capacities

    #Create one big list of all indexes in order to know which one is the next
    all_indxs = Chg_indxs + DChg_indxs
    all_indxs.sort()
    # Iterating over all indexes of charge start
    for i, indx in enumerate(Chg_indxs):
        i_all = all_indxs.index(indx) #Finding the corresponding item in list of all indexes
        start = indx   #Setting start of Charge to the charge index
        try:            #Setting end of charge to discharge index, or last index of data.
            end = all_indxs[i_all+1]
        except:
            end = len(data)-1
        #Getting the data slice, and putting it in charges list.
        voltages, capacities = slice_data(start, end, data)
        #print(voltages)
        charges.append((voltages, capacities))
    
    # Iterating over all indexes of discharge start
    for i, indx in enumerate(DChg_indxs):
        i_all = all_indxs.index(indx) #Finding the corresponding item in list of all indexes
        start = indx   #Setting start of Discharge to the discharge index
        try:            #Setting end of charge to charge index, or last index of data.
            end = all_indxs[i_all+1]
        except:
            end = len(data)-1
        #Getting the data slice, and putting it in charges list.
        voltages, capacities = slice_data(start, end, data)
        discharges.append((voltages, capacities))

            
    return charges, discharges"""


def read_xlsx(filename):
    import openpyxl as px
    import numpy as np

    # Open workbook
    wb = px.load_workbook(filename)
    # Get sheet
    sheet = wb[wb.sheetnames[3]]
    max_row = sheet.max_row-1 #-1 removes the first row with text
    # Spawn arrays for all raw data of columns
    status = []
    voltage = np.zeros(max_row)
    capacity = np.zeros(max_row)

    # Insert all data in arrays
    for i in range(max_row):
        voltage[i] = sheet.cell(row=i+2, column = 7).value
        capacity[i] = sheet.cell(row=i+2, column = 8).value
        status.append(sheet.cell(row=i+2, column=2).value)

    # Find where a charge and a discharge starts. Since charging occurs first, charges will be at even indexes of indx, eg: 0, 2, 4..
    indx =[] # Just a list to store where in the raw data a charge (even index) or discharge (odd index)
    for i in range(1,len(status)):
        if status[i] == "CC Chg" and status[i-1] != "CC Chg": # If status just changed to CC Chg, this is a charge start
            indx.append(i)
        elif status[i] == "CC DChg" and status[i-1] != "CC DChg": # If status just changed to CC DChg, this is a discharge start
            indx.append(i)
    indx.append(max_row) #Inserting last elem means we catch the end cycle!

    # Making lists for the cycles
    charges = []
    discharges = []

    # Inserting slices of raw data (corresponding to 1 cycle) into the list of charges
    for i in range(len(indx)-1): #-1 is to not run out of indexes when i+1
        if (i % 2) == 0: #index is even -> charge cycle
            charges.append((voltage[indx[i]:indx[i+1]], capacity[indx[i]:indx[i+1]])) #Inserting tuple of two arrays sliced to be one cycle of voltages and capacities.
        else: #Not even? Then its a discharge cycle
            discharges.append((voltage[indx[i]:indx[i+1]], capacity[indx[i]:indx[i+1]]))

    return charges, discharges